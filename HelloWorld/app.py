# print("HeloWorld")
# name = input('input your name: ')
# print('name: ' + name)
# # --------------------------- #
# x, y, z = [1, 2, 3]
print('# ----------------------')

from ecommerce.shipping import calc_shipping
calc_shipping()
print('# ----------------------')

import random
for i in range(3):
    print(random.random())

for i in range(3):
    print(random.randint(10,20))

print('# ----------------------')

members = ['Bob', 'John', 'Ava']
leader = random.choice(members)
print(leader)
print('# ----------------------')


class Dice:
    def calc(self):
        import random
        first = random.randint(1,6)
        second = random.randint(1,6)
        return first, second


dice = Dice()
print(dice.calc())
print('# ----------------------')

# PATH AND DIRECTORY
print('# -------------------- #')

print(" PATHLIB PACKAGE")
from pathlib import Path

#ABSOLUTE PATH; (c:\programfiles\windows)
#RELATIVE PATH(...)

path = Path("ecommerce")
# path2 = Path("email")
# path2.mkdir()
# path2.rmdir()
print(path.exists())
print('# ----------------------')

path = Path()
print(path.glob("*.*")) #'*', ...other pattern
for file in path.glob("*"): # try '*.py'
    print(file)
print('# ----------------------')

#USE EXEL FILES, DOWNMLOAD OPENPYXL PACKAGE THEN IMPORT IT
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1, 1)
# print(cell.value)
print(sheet.max_row) # num of rows
print('# ----------------------')

for row in range(1, sheet.max_row + 1):
    print(row)

print('# ----------------------')

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value, end =' ')

print('# ----------------------')

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')